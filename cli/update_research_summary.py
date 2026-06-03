"""Maintain the TrueKnot research summary workbook (Phase 9).

Single reusable entry point the OpenClaw `research-summary` skill calls to keep
`<TK Research>/final/TrueKnot-Research-Summary.xlsx` current, where <TK Research>
is the trueknotsg Google Drive My Drive folder (see _tk_base / TK_RESEARCH_BASE).
NOTE: superseded going forward by the native Google Sheet maintained via the
`update-summary` skill (gog); kept for the xlsx artifact.

  1. consolidate every report PDF into final/pdf/research-<date>-<ticker>.pdf
  2. sync the A+ rows from REGISTER.md (authoritative rating / price / EV)
  3. refresh "Current Price" + "Move Since Report" via yfinance latest close
  4. colour-code by rating, link each row to its PDF, scrub architecture leaks

Design notes
------------
- Columns are addressed BY HEADER NAME, never by fixed index — so a column the
  operator/agent adds or reorders is preserved and still updated correctly.
- The EXISTING summary in final/ is the base when present (preserves manual
  notes + any added columns); the Downloads template is the first-run fallback.
- Idempotent: safe to run repeatedly.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SHEET = "Research Summary"
RATING_FILL = {"Overweight": "C6EFCE", "Hold": "FFEB9C", "Underweight": "FFC7CE", "Stale": "BFBFBF"}
# Canonical columns the updater manages. Operator-added columns are left alone.
CANON_HEADERS = [
    "Report Date", "Ticker", "Rating", "Price at Report ($)", "EV 12-Month ($)",
    "EV 12-Month (%)", "Current Price (latest close)", "Move Since Report (%)",
    "Notes", "Report PDF",
]


def parse_register(register_path: Path) -> dict[str, dict]:
    """Parse REGISTER.md's A+ table → {ticker: {date,rating,ref,ev,evpct}}."""
    out: dict[str, dict] = {}
    if not register_path.exists():
        return out
    for line in register_path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not re.match(r"\| 2026-\d\d-\d\d ", s):
            continue
        c = [x.strip() for x in s.strip("|").split("|")]
        date, tkr, rating, grade, ref, ev, evs = (c + [""] * 7)[:7]
        try:
            out[tkr] = {
                "date": date, "rating": rating.title(),
                "ref": float(re.sub(r"[^0-9.]", "", ref)),
                "ev": float(re.sub(r"[^0-9.]", "", ev)),
                "evpct": float(re.sub(r"[^0-9.\-]", "", evs)) / 100,
            }
        except ValueError:
            continue
    return out


def _week_dir_for_date(date_str: str) -> str:
    """ISO-calendar week folder for a run date, e.g. '2026-05-26' → 'wk 22 2026'.
    Final reports are organised under final/<week>/<run>/ + final/<week>/pdf/."""
    try:
        iso = datetime.strptime(date_str, "%Y-%m-%d").isocalendar()
        return f"wk {iso[1]} {iso[0]}"
    except (ValueError, TypeError):
        return ""


def consolidate_pdfs(final_dir: Path) -> int:
    """Move report PDFs from final/<week>/<run>/ into that week's pdf/ folder.
    Week-aware (2026-06-02): reports live under final/wk NN YYYY/<date>-<ticker>/.
    Returns count moved. Also handles legacy final/<run>/ layout."""
    moved = 0
    # week-organised layout: final/wk NN YYYY/<date>-<ticker>/research-*.pdf
    for run in final_dir.glob("wk */2026-*"):
        if not run.is_dir():
            continue
        pdfdir = run.parent / "pdf"
        pdfdir.mkdir(parents=True, exist_ok=True)
        for pdf in run.glob("research-*.pdf"):
            shutil.move(str(pdf), str(pdfdir / pdf.name))
            moved += 1
    # legacy flat layout: final/<date>-<ticker>/
    for run in final_dir.glob("2026-*"):
        if not run.is_dir():
            continue
        pdfdir = final_dir / "pdf"
        pdfdir.mkdir(parents=True, exist_ok=True)
        for pdf in run.glob("research-*.pdf"):
            shutil.move(str(pdf), str(pdfdir / pdf.name))
            moved += 1
    return moved


def _yf_last_close(ticker: str) -> float | None:
    try:
        import yfinance as yf
        from tradingagents.dataflows.stockstats_utils import yf_retry
        h = yf_retry(lambda: yf.Ticker(ticker).history(period="7d", auto_adjust=False))
        if h is None or h.empty:
            return None
        return round(float(h["Close"].iloc[-1]), 2)
    except Exception:
        return None


def _header_map(ws) -> tuple[int, dict[str, int]]:
    """Find the header row (the one containing 'Ticker') → (row, {name: col})."""
    for r in range(1, min(ws.max_row, 6) + 1):
        names = {str(ws.cell(row=r, column=c).value).strip(): c
                 for c in range(1, ws.max_column + 1) if ws.cell(row=r, column=c).value}
        if "Ticker" in names:
            return r, names
    raise ValueError("could not locate a header row containing 'Ticker'")


def _strip_marker(note: str | None) -> str | None:
    if not note:
        return note
    note = re.sub(r"\s*[—-]\s*REFRESH QUEUED", "", note)
    note = re.sub(r"STALE \+?\d+%\s*\|?\s*", "", note)
    return note.strip(" |")


def update_summary(base_path: Path, register_path: Path, final_dir: Path,
                   out_path: Path, price_fn=_yf_last_close) -> dict:
    aplus = parse_register(register_path)

    if base_path.exists():
        wb = load_workbook(base_path)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    else:  # first-run fallback: minimal fresh workbook
        wb = Workbook(); ws = wb.active; ws.title = SHEET
        for j, h in enumerate(CANON_HEADERS, start=1):
            ws.cell(row=1, column=j, value=h).font = Font(bold=True)

    hrow, cols = _header_map(ws)
    # ensure canonical columns exist (append any missing at the end)
    for h in CANON_HEADERS:
        if h not in cols:
            c = ws.max_column + 1
            ws.cell(row=hrow, column=c, value=h).font = Font(bold=True)
            cols[h] = c

    def col(name):
        return cols[name]

    # index existing data rows by ticker
    row_of: dict[str, int] = {}
    for r in range(hrow + 1, ws.max_row + 1):
        t = ws.cell(row=r, column=col("Ticker")).value
        if t:
            row_of[str(t).strip()] = r

    # sync A+ rows (update existing or append new)
    next_row = ws.max_row + 1
    for tkr, d in aplus.items():
        r = row_of.get(tkr)
        if r is None:
            r = next_row; next_row += 1; row_of[tkr] = r
            ws.cell(row=r, column=col("Ticker"), value=tkr).font = Font(bold=True)
            ws.cell(row=r, column=col("Notes"), value=f"A+ {d['date']} - see report")
        ws.cell(row=r, column=col("Report Date"), value=d["date"])
        rc = ws.cell(row=r, column=col("Rating"), value=d["rating"])
        rc.fill = PatternFill("solid", fgColor=RATING_FILL.get(d["rating"], "FFFFFF"))
        ws.cell(row=r, column=col("Price at Report ($)"), value=d["ref"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=col("EV 12-Month ($)"), value=d["ev"]).number_format = "$#,##0.00"
        ws.cell(row=r, column=col("EV 12-Month (%)"), value=d["evpct"]).number_format = "+0.00%;-0.00%"
        note = _strip_marker(ws.cell(row=r, column=col("Notes")).value)
        if tkr in ("AAPL", "FUTU", "ORCL"):
            note = f"A+ refresh {d['date']} - see report"
        ws.cell(row=r, column=col("Notes"), value=note)
        fn = f"research-{d['date']}-{tkr}.pdf"
        pcell = ws.cell(row=r, column=col("Report PDF"), value=fn)
        # Week-aware PDF location: final/<week>/pdf/<fn>. Fall back to the
        # legacy final/pdf/<fn> for any pre-week-layout reports.
        week = _week_dir_for_date(d["date"])
        if week and (final_dir / week / "pdf" / fn).exists():
            pcell.hyperlink = f"{week}/pdf/{fn}"
            pcell.font = Font(color="0563C1", underline="single", size=9)
        elif (final_dir / "pdf" / fn).exists():
            pcell.hyperlink = f"pdf/{fn}"
            pcell.font = Font(color="0563C1", underline="single", size=9)

    # refresh current price + move for ALL rows
    for tkr, r in row_of.items():
        cur = price_fn(tkr)
        if cur is None:
            continue
        ws.cell(row=r, column=col("Current Price (latest close)"), value=cur).number_format = "$#,##0.00"
        ref = ws.cell(row=r, column=col("Price at Report ($)")).value
        if isinstance(ref, (int, float)) and ref:
            mc = ws.cell(row=r, column=col("Move Since Report (%)"), value=(cur - ref) / ref)
            mc.number_format = "+0.00%;-0.00%"

    # tidy borders
    thin = Side(style="thin", color="D9D9D9"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(hrow, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = bd

    # scrub architecture leak in any Legend sheet
    if "Legend" in wb.sheetnames:
        for row in wb["Legend"].iter_rows():
            for cell in row:
                if cell.value and "Multi-Agent Research Pipeline" in str(cell.value):
                    cell.value = "TrueKnot Equity Research"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"out": str(out_path), "rows": ws.max_row - hrow, "aplus": len(aplus)}


def _tk_base() -> Path:
    """TK Research base — mirrors cli.research._tk_research_base (kept in sync):
    the trueknotsg Google Drive My Drive folder, overridable via TK_RESEARCH_BASE."""
    override = os.environ.get("TK_RESEARCH_BASE")
    if override:
        return Path(override).expanduser()
    return (
        Path.home()
        / "Library" / "CloudStorage"
        / "GoogleDrive-trueknotsg@gmail.com" / "My Drive" / "TK Research"
    )


def main(argv=None) -> int:
    tk = _tk_base()
    p = argparse.ArgumentParser(prog="update_research_summary")
    p.add_argument("--tk-dir", default=str(tk))
    p.add_argument("--register", default=None)
    p.add_argument("--out", default=None)
    p.add_argument("--no-pdf-consolidate", action="store_true")
    a = p.parse_args(argv)
    tk = Path(a.tk_dir)
    final_dir = tk / "final"
    out = Path(a.out) if a.out else final_dir / "TrueKnot-Research-Summary.xlsx"
    register = Path(a.register) if a.register else tk / "REGISTER.md"
    moved = 0 if a.no_pdf_consolidate else consolidate_pdfs(final_dir)
    res = update_summary(out, register, final_dir, out)
    print(f"summary: {res['out']} | rows={res['rows']} A+={res['aplus']} pdfs_consolidated={moved}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
