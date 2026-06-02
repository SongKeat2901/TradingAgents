"""Tests for the research-summary updater (Phase 9)."""

import pytest
from openpyxl import Workbook, load_workbook

pytestmark = pytest.mark.unit

_REGISTER = """# register
| Trade date | Ticker | Rating | Grade | Ref price | 12-mo EV | EV vs spot | QC | PDF |
|---|---|---|---|---|---|---|---|---|
| 2026-05-29 | INTC | Underweight | A+ | $120.89 | $124.03 | +2.60% | pass | x.pdf |
| 2026-05-26 | GOOGL | HOLD | A+ | $382.97 | $393.25 | +2.68% | pass | x.pdf |
"""


def _setup(tmp_path):
    tk = tmp_path / "TK Research"
    final = tk / "final"; (final / "pdf").mkdir(parents=True)
    (tk / "REGISTER.md").write_text(_REGISTER, encoding="utf-8")
    # PDFs present so the link is set
    (final / "pdf" / "research-2026-05-29-INTC.pdf").write_text("x")
    (final / "pdf" / "research-2026-05-26-GOOGL.pdf").write_text("x")
    return tk, final


def test_parse_register(tmp_path):
    from cli.update_research_summary import parse_register
    tk, _ = _setup(tmp_path)
    d = parse_register(tk / "REGISTER.md")
    assert set(d) == {"INTC", "GOOGL"}
    assert d["GOOGL"]["rating"] == "Hold"  # normalized from HOLD
    assert d["INTC"]["ref"] == 120.89 and round(d["INTC"]["evpct"], 4) == 0.026


def test_first_run_builds_and_syncs(tmp_path):
    from cli.update_research_summary import update_summary
    tk, final = _setup(tmp_path)
    out = final / "TrueKnot-Research-Summary.xlsx"
    update_summary(out, tk / "REGISTER.md", final, out, price_fn=lambda t: 100.0)
    wb = load_workbook(out); ws = wb.active
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Ticker" in hdr and "Report PDF" in hdr
    tickers = {ws.cell(row=r, column=hdr.index("Ticker") + 1).value for r in range(2, ws.max_row + 1)}
    assert {"INTC", "GOOGL"} <= tickers
    # GOOGL rating normalized + colored
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=hdr.index("Ticker") + 1).value == "GOOGL":
            assert ws.cell(row=r, column=hdr.index("Rating") + 1).value == "Hold"
            assert ws.cell(row=r, column=hdr.index("Report PDF") + 1).hyperlink is not None


def test_operator_added_column_is_preserved(tmp_path):
    """A column the operator added (by header) must survive an update."""
    from cli.update_research_summary import update_summary
    tk, final = _setup(tmp_path)
    out = final / "TrueKnot-Research-Summary.xlsx"
    # build a base with INTC + a custom "Conviction" column with a value
    wb = Workbook(); ws = wb.active; ws.title = "Research Summary"
    headers = ["Report Date", "Ticker", "Rating", "Price at Report ($)", "EV 12-Month ($)",
               "EV 12-Month (%)", "Current Price (latest close)", "Move Since Report (%)",
               "Notes", "Report PDF", "Conviction"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    ws.cell(row=2, column=2, value="INTC")
    ws.cell(row=2, column=11, value="High")  # Conviction
    wb.save(out)

    update_summary(out, tk / "REGISTER.md", final, out, price_fn=lambda t: 130.0)
    wb2 = load_workbook(out); ws2 = wb2.active
    hdr = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
    assert "Conviction" in hdr  # custom column preserved
    crow = next(r for r in range(2, ws2.max_row + 1)
                if ws2.cell(row=r, column=hdr.index("Ticker") + 1).value == "INTC")
    assert ws2.cell(row=crow, column=hdr.index("Conviction") + 1).value == "High"  # value preserved
    # and INTC's managed columns got synced
    assert ws2.cell(row=crow, column=hdr.index("EV 12-Month ($)") + 1).value == 124.03


def test_week_dir_helper():
    from cli.update_research_summary import _week_dir_for_date
    assert _week_dir_for_date("2026-05-26") == "wk 22 2026"
    assert _week_dir_for_date("2026-05-29") == "wk 22 2026"
    assert _week_dir_for_date("bad") == ""


def test_consolidate_pdfs_week_layout(tmp_path):
    from cli.update_research_summary import consolidate_pdfs
    final = tmp_path / "final"
    run = final / "wk 22 2026" / "2026-05-29-INTC"
    run.mkdir(parents=True)
    (run / "research-2026-05-29-INTC.pdf").write_text("x")
    moved = consolidate_pdfs(final)
    assert moved == 1
    assert (final / "wk 22 2026" / "pdf" / "research-2026-05-29-INTC.pdf").exists()
    assert not (run / "research-2026-05-29-INTC.pdf").exists()


def test_pdf_link_is_week_relative(tmp_path):
    from cli.update_research_summary import update_summary
    tk = tmp_path / "TK Research"; tk.mkdir()
    final = tk / "final"
    wkpdf = final / "wk 22 2026" / "pdf"; wkpdf.mkdir(parents=True)
    (wkpdf / "research-2026-05-29-INTC.pdf").write_text("x")
    reg = tk / "REGISTER.md"; reg.write_text(_REGISTER)
    out = final / "TrueKnot-Research-Summary.xlsx"
    update_summary(out, reg, final, out, price_fn=lambda t: None)
    wb = load_workbook(out); ws = wb["Research Summary"]
    # find INTC row's PDF cell hyperlink
    links = [c.hyperlink.target for row in ws.iter_rows() for c in row
             if c.hyperlink and "INTC" in str(c.value)]
    assert links and links[0] == "wk 22 2026/pdf/research-2026-05-29-INTC.pdf"
